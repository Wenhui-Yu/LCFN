import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook

def print_params(para_name, para):
    for i in range(len(para)):
        print(para_name[i]+':  ',para[i])

def print_value(value):
    [inter, loss, f1_max, F1, NDCG] = value
    print('iter: %d loss %.2f f1 %.4f' %(inter, loss, f1_max), end='  ')
    print(F1, NDCG)

def save_params(para_name,para,path_excel):
    wb = Workbook( )
    table = wb.active
    table.title = 'Parameters'
    ldata = []
    for i in range(1, len(para_name)):  # do not save GPU_index
        parameter = [para_name[i]]
        parameter_value = para[i]
        if isinstance(parameter_value, list):
            for value in parameter_value:
                parameter.append(value)
        elif isinstance(parameter_value, bool): parameter.append({True: 'Yes', False: 'No'}[parameter_value])
        else: parameter.append(parameter_value)
        ldata.append(parameter)
    for i, p in enumerate(ldata):
        for j, q in enumerate(p):
            table.cell(row = i+1, column = j+1).value = q
    wb.save(path_excel)
    wb.close()

def save_value(df_list,path_excel,first_sheet):
    excelWriter = pd.ExcelWriter(path_excel, engine='openpyxl',mode='a')

    if first_sheet is False:
        workbook = load_workbook(path_excel)
        excelWriter.book = workbook
        exist_sheets = workbook.get_sheet_names()
        for df in df_list:
            if df[1] in exist_sheets:
                workbook.remove_sheet(workbook.get_sheet_by_name(df[1]))
            df[0].to_excel(excel_writer=excelWriter, sheet_name=df[1],index = True)
            excelWriter.save()
    else:
        for df in df_list:
            df[0].to_excel(excel_writer=excelWriter, sheet_name=df[1], index=True)
            excelWriter.save()
    excelWriter.close()

def df2str(df):
    df_str = ''
    for i in range(df.shape[0]):
        df_list = df.iloc[[i], :].values.tolist()
        df_list2 = [str(i) for i in df_list]
        str_temp = ''.join(df_list2)
        df_str = df_str +str_temp+','
    return df_str