#-*- coding:utf-8 –*-
## Show the performances varied with one or two parameter(s)
## Huidi Zhang 2018.11.20
## author @Huidi Zhang, zhd16@mails.tsinghua.edu.cn

import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
import time
from print_save import save_value
from print_save import df2str
import os
from numpy import *
import sys
import operator

k = 2           # 这个参数用于标注如果在一组文件中有两个参数发生改变，则输出的表中选top k的那个值

def read_data_from_sheet(path, sheet, colu, ind):
    df = pd.DataFrame(pd.read_excel(path, sheetname=sheet, header=0, index_col=0))
    if colu >= 0:
        data = df.loc[ind, colu]
    else:
        data = df.loc[ind, ]
    return data

path = os.path.abspath(os.path.dirname(os.getcwd())) + '/experiment_result'
path_read = path + '/data_process'

file_dict = dict()                # 对所有的文件简历字典，key=model+dataset，value为key相同的文件的集合

if len(os.listdir(path_read)) == 0:
    print("( ˃ ˄ ˂̥̥ )     ", path_read+" is empty!!")
    sys.exit(0)

for file_name in os.listdir(path_read):     # 读入所有等待被处理的文件

    if operator.eq(file_name[-4:], 'xlsx') == 1:    # 判断这个文件是不是xlsx文件，如果是，则进行下面的操作（通过这个步骤过滤掉其他后缀的文件以及该目录下的文件夹）
        file_p = path_read + '/' + file_name  # 待处理的xlsx文件的路径
        parameter_df = pd.DataFrame(pd.read_excel(file_p, sheetname=0, index_col=0)) # 读入待处理文件的第一个sheet（这个sheet里存储着实验参数）
        model = parameter_df.loc['MODEL',1]   # 取出model和dataset的值
        dataset = parameter_df.loc['DATASET',1]
        model_dataset = str(dataset)+'_'+str(model) # 用model和dataset组成一个字符串，作为file_dict的key
        if file_dict.get(model_dataset) is None:    # 如果这个key不存在，则建一个新的
            file_set = set()
        else:
            file_set = file_dict.get(model_dataset)   # 如果这个key存在，则获取这个key
        file_set.add(file_name)                       # 将这个文件名加入set
        file_dict.update({model_dataset: file_set})   # 更新字典

for model_dataset, file_set in file_dict.items():    # 对于每一组model-dataset进行处理
    para_dict = dict()                                # 用字典存储这个model和dataset的组合的实验中出现过的参数，key为参数名称，value为包含所有出现过的参数值的set
    list_index = []                                   # 存储parameter的index，也就是所有的参数名
    for file_name in file_set:                       # 对所有dataset和model相同的实验结果文件进行处理
        file_p = path_read + '/' + file_name         # 生成文件路径
        sheets = load_workbook(file_p).sheetnames     # 获取这个文件的所有sheet名，以备后用
        parameter_df = pd.DataFrame(pd.read_excel(file_p, sheetname=0, index_col=0))  # 获得这个文件的参数dataframe
        list_index = parameter_df.index               # 获得这个文件的所有参数名
        for index in list_index:                      # 循环处理每一个参数
            if para_dict.get(index) is None:          # 如果这个参数在参数dict里还不存在，则新建一个存着所有参数值的set
                para_set = set()
            else:
                para_set = para_dict.get(index)
            if index == 'TOP_K':                      # 如果这个参数是top_k，则把top_k的所有值都取出来，转成str存储
                value = df2str( pd.DataFrame(parameter_df.loc[index, :]).T)
            else:
                value = str(parameter_df.loc[index, 1])  # 如果参数不是top-k,则只取第一位就可以了
                if value.isdigit():
                    value = float(value)
            para_set.add(value)                          # 给这个参数对应的参数值的set添加上这个新值
            para_dict.update({index:para_set})           # 更新字典

    para_df = pd.DataFrame()   # 用来存储每个index的名字以及其对应的参数值（所有出现过的参数值）
    changed_para = []          # 存储所有值发生过改变的参数

    for key in list_index:
        value = para_dict.get(key)
        if len(value)>1:             # 如果这个参数对应的value（是个set）里存的值多于一个
            changed_para.append(key)  # 说明这个参数的值在不同次的实验中发生了改变
        temp_list = []
        temp_list.append(key)
        for v in value:
            temp_list.append(v)         # 将参数名和参数值连接成一个list
        para_df = para_df.append(pd.DataFrame(temp_list).T)  # 将这个list存成dataframe

    para_df = para_df.set_index(0)  # 将第0列设为index（第0列是参数名）
    changed_para_str = ', '.join(changed_para)  # 将改变了的参数的参数名连成字符串，用于生产输出文件的文件名
    path_write = path + '/data_collection'+ '/' + model_dataset + '_' +changed_para_str+'_'+str(int(time.time())) + str(int(random.uniform(100, 900))) + '.xlsx'
    save_value([[para_df, 'Parameters']], path_write, first_sheet=True)  # 将所有出现过的参数名和参数值存下来

    if len(changed_para) == 1 or len(changed_para) == 0:  # 只有一个参数发生改变(没有参数改变的也当做只有一个参数改变的特殊情况)
        if len(changed_para) == 1 :
            index_name = str(str(changed_para[0]))
            print("(*/ω＼*)     ", model_dataset + ": " + changed_para[0] + " is the variable")  # 输出发生改变的参数
        else:
            print("o(≧口≦)o    ", model_dataset + ": there is no changed parameters")
            index_name = "DATASET"
        top_k = str(para_df.loc['TOP_K', 1])  # 对top_k的值进行一些字符串处理
        top_k = top_k.strip(',')
        top_k = top_k.strip('[')
        top_k = top_k.strip(']')

        for sheet in sheets:
            if operator.eq(sheet, 'Parameters') == 0 and operator.eq(sheet, 'Filename') == 0:
                output_df = pd.DataFrame(columns=[int(float(x)) for x in str(top_k).split(',')])
                output_df.columns.name = 'TOP_K'
                output_df.index.name = index_name
                for file_name in file_set:
                    if operator.eq(file_name[-4:], 'xlsx') == 1:
                        file_p = path_read + '/' + file_name
                        parameter_df = pd.DataFrame(pd.read_excel(file_p, sheetname=0, index_col=0))
                        p = parameter_df.loc[index_name, 1]
                        output_df.loc[p] = read_data_from_sheet(file_p, sheet, colu=-1, ind='mean').T
                output_df = output_df.sort_index()
                output_df.dropna(axis=0, how='all', inplace=True)
                output_df.dropna(axis=1, how='all', inplace=True)
                save_value([[output_df, sheet]], path_write, first_sheet=False)
    elif len(changed_para) == 2: # 有两个参数发生改变
        print("(๑‾ ꇴ ‾๑)     ", model_dataset + ": " + changed_para[0] + " and " + changed_para[1] + " are two variables") # 输出发生改变的参数
        for sheet in sheets:
            if operator.eq(sheet, 'Parameters') == 0 and operator.eq(sheet, 'Filename') == 0:
                output_df = pd.DataFrame()
                output_df.index.name = changed_para[1] # 设置dataframe的行名和列名
                output_df.columns.name = changed_para[0]
                for file_name in file_set:
                    if operator.eq(file_name[-4:], 'xlsx') == 1:
                        file_p = path_read + '/' + file_name
                        parameter_df = pd.DataFrame(pd.read_excel(file_p, sheetname=0, index_col=0))
                        p = parameter_df.loc[changed_para[0], 1]
                        q = parameter_df.loc[changed_para[1], 1]
                        output_df.loc[q,p] = read_data_from_sheet(file_p, sheet, colu=k, ind='mean').T
                output_df = output_df.sort_index()  # 对index排序
                col = output_df.columns.values.tolist()    # 对column排序
                col.sort()
                output_df = output_df.ix[:,col]
                output_df.dropna(axis=0, how='all', inplace= True)
                output_df.dropna(axis=1, how='all', inplace=True)
                save_value([[output_df, sheet]], path_write, first_sheet=False)

    else:  # 超过两个参数发生改变
        print( "(๑‾  ‾๑)     ", model_dataset + ":more than two parameters are changed")

files = os.listdir(path_read)  # 删除dataprocess里的所有文件
for file in files:
    os.remove(path_read + '/' + file)
