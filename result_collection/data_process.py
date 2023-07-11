#-*- coding:utf-8 –*-
## Processing the experiment data, collecting data with the same parameter setting, and calculating aver and std.
## Huidi Zhang 2018.11.20
## author @Huidi Zhang, zhd16@mails.tsinghua.edu.cn

import pandas as pd
from openpyxl import load_workbook #@x
from openpyxl import Workbook #@x
import time
from print_save import save_value #@x
from print_save import df2str #@x
import os
from numpy import *
import operator

top_ave = 5                                 # 设置实验结果要取所有迭代的前多少轮的平均值

path = os.path.abspath(os.path.dirname(os.getcwd())) + '/experiment_result'
emoji = ["(๑‾  ‾๑)","( ˃ ˄ ˂̥̥ ) ","(*/ω＼*)","o(≧口≦)o","╰(*°▽°*)╯","~( ﹁ ﹁ ) ~~","(⊙﹏⊙)","(＠_＠;)"]

def process_metric(df, method, para):       # input：df是指要被处理的dataframe；method有两种选项：max和top；para是method=top时要用到的参数
    output_list = []
    if operator.eq(method, 'top') == 1:             # 取这个metric的每一列的top_para 的平均值
        if para > df.shape[0]:              # 判断para是不是比metric的行数多
            para = df.shape[0]

        def top_average(column):            # 这个函数定义了如何求每一列top_para的平均值
            num_sum = 0
            new_column = column.sort_values(ascending=False)    # 将这一列降序排序
            for i in range(para):                               # 将前para个数相加
                num_sum = num_sum + new_column.iloc[i]
            return num_sum / para                               # 求平均
        output_list = df.apply(top_average)                     # 对metric的每一列求top_para的平均值
    else:
        if operator.eq(method, 'max') == 1:
            output_list = df.apply(lambda x: x.max())
    return output_list                              # output：输出一个处理后的list


path_read = path                                    # 从该路径读取要被处理的文件
file_dict = dict()                                  # 通过字典存储参数相同的文件，key：所有参数拼接得到的字符串，value:该参数对应的若干个文件

for file_name in os.listdir(path_read):             # 读入所有等待被处理的文件
    if operator.eq(file_name[-5:], '.xlsx') == 1:           # 判断这个文件是不是xlsx文件，如果是，则进行下面的操作（通过这个步骤过滤掉其他后缀的文件以及该目录下的文件夹）
        file_p = path_read + '/' + file_name       # 待处理的xlsx文件的路径
        parameter_df = pd.DataFrame(pd.read_excel(file_p, sheetname=0,header = None))  # 读入待处理文件的第一个sheet（这个sheet里存储着实验参数），header=None表明在这个表格中并没有表头
        parameter_str = df2str(parameter_df)        # 将实验参数的dataframe转为字符串，作为字典的key
        if file_dict.get(parameter_str) is None:    # 如果字典中这个key还不存在
            file_list = []                          # 新建一个file list，这个list里存储这个key对应的所有文件
        else:
            file_list = file_dict.get(parameter_str)    # 如果这个key已经存在，则获取key对应的文件list
        file_list.append(file_name)                     # 把刚刚读入的文件加入到list里
        file_dict.update({parameter_str: file_list})    # 更新字典

for key, value in file_dict.items():                    # dict.items方法会将键-值对作为元组返回
    sheets = set()                                      # 把参数一样的这几个文件里出现过的sheets都记在set里，防止某个实验跑的时候没生成好，没有把结果存下来（“6k异常问题的解决”）
    for n in value:
        #wb = load_workbook(path_read + '/' + value[0])     # 加载这个key 所对应的文件列表中的第一个文件（因为这个列表里的文件的参数都一样，所以取第一个文件的参数就可以）
        wb = load_workbook(path_read + '/' + n)
        #sheets = wb.sheetnames                              # 保留文件的sheets，以备后用
        for e in wb.sheetnames:
            sheets.add(e)
    sheets = list(sheets)
    sheets.sort()

    parameter = pd.DataFrame(pd.read_excel(path_read + '/' + value[0], sheetname=0, header = None , index_col=0)) # 将文件的参数存成dataframe的形式，将第0列设置为index
    parameter.index.name = 'para'                       # 设置dataframe的行名和列名
    parameter.columns.name = 'value'

    dataset = str(parameter.loc['DATASET',1])       # 保留文件的dataset和model，作为后面输出文件的文件名
    model = str(parameter.loc['MODEL',1])
    #eta = str(parameter.loc['eta',1])
    #lambda_r = str(parameter.loc['lambda_r',1])
    path_write = path + '/data_process/' + dataset + '_' + model + '_' + str(int(time.time())) + str(int(random.uniform(100, 900))) + '.xlsx'  # 输出文件的文件路径
    #path_write = path + '/data_process/' +dataset+'_'+model+'_eta='+eta+'_lambda='+lambda_r+'_' + str(int(time.time())) + str(int(random.uniform(100, 900))) + '.xlsx'
    save_value([[parameter, 'Parameters']], path_write, first_sheet=True)  # 将参数存进excel
    for sheet in sheets:                                                     # 处理表中的每一个sheet，包括F1和NDCG
        if operator.eq(sheet, 'Parameters') == 0 and operator.eq(sheet, 'Filename') == 0: # 如果这个sheet不是parameter也不是filename，则进行下面的操作
            df_max = pd.DataFrame()                                          # 存储这个sheet（F1或者NDCG）对应的处理结果，df_max表示这个表里的结果是之前待处理的表的max值合成的
            df_top = pd.DataFrame()                                          # df_top表示这个表里的结果是之前待处理的表的top_ave的平均值合成的
            for file_p in value:                                               # 对这个key对应的value的list里的每一个文件进行处理
                temp_f = load_workbook(path_read + '/' + file_p)
                temp_sn = temp_f.sheetnames
                if sheet in temp_sn:
                    metric = pd.DataFrame(pd.read_excel(path_read + '/' + file_p, sheetname=sheet, header=0, index_col=0)) # 读入某一个文件里的一个sheet
                    list_max = process_metric(metric,method = 'max',para = top_ave  )             # 处理这个metric，得到F1_max或NDCG_max,得到的值为一行
                    list_top = process_metric(metric, method = 'top', para = top_ave )            # 处理这个metric，得到F1_top或NDCG_top，得到的值为一行
                    df_max = df_max.append(list_max, ignore_index=True)           # 将list_max 和 list_top分别加在对应的dataframe上
                    df_top = df_top.append(list_top, ignore_index=True)           # 经过上述操作，在df_top 和df_max里，每一行由一个实验结果文件生成
            df_top = df_top.append(df_top.mean(),ignore_index=True)           # 再将上述所有各次实验的平均值追加在对应的dataframe后面
            df_max = df_max.append(df_max.mean(), ignore_index=True)
            df_top = df_top.append(df_top.std(), ignore_index=True)  # 再将上述所有各次实验的方差追加在对应的dataframe后面
            df_max = df_max.append(df_max.std(), ignore_index=True)
            index_list = list(range(1,df_top.shape[0]-1))                         # index_list作为df_top 和df_max的index
            index_list.append('mean')                                        # index的格式为：最后一行的index是“mean”，之前的index是从1 开始的数字，表示第几次实验
            index_list.append('std')
            df_top['epoch'] = index_list                                     # 将生成的index赋值给epoch这一列

            df_top = df_top.set_index('epoch')                               # 把epoch这一列设置为index
            df_max['epoch'] = index_list
            df_max = df_max.set_index('epoch')
            save_value([[df_max, sheet+'_max'],[df_top,sheet + '_top'+str(top_ave)]], path_write, first_sheet=False) # 将df_top 和df_max写入文件

    print(value, "   "+emoji[random.randint(0,len(emoji))])
    file_dataframe = pd.DataFrame(value)  # 将此次处理的所有文件（生成这个表格所用到的文件）的文件名记录下来，写入到最后一个sheet里
    save_value([[file_dataframe, 'Filename']], path_write, first_sheet=False)